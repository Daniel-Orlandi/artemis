import datetime
import os
from dataclasses import dataclass

from artemis.utils import check_exists, check_d_type
from artemis.utils.data_logger import Logger
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation


class WorkBook:
    """
        class to provide methods to write data to excel files         

    Attributes
    ----------
    __data_workbook = None
        openpyxl object representing excel file.

    __active_table = None
        openpyxl object representing selected table in excel file.          

    self.__data_frame: pandas.DataFrame
        dataframe where data will be stored after processing.

    self.__location_list: list of Location objects
        data from dataframe will be location-wise
        selected and stored in location objects

    self.logger: utils.data_logger
        logger initializator

    self.todays_date 
        
    Returns
    ----------
    None

    Methods
    ----------
    get_workbook(self):

    get_active_table(self):

    get_location_list(self):

    to_xlsx(workbook, dataframe, data_type: str):

    set_date_cell(self, date: str, pos_in_table: str):

    set_cell(self, data_list :list, pos_in_table_list: list):

    set_active_table(self, table_name: str):

    write_data_to_table(self):

    load_sheet(self, **kwargs) -> None:

    save_sheet(self, workbook, out_file: str):   

    """
    def __init__(self, filename: str, location_list: list) -> None:
        self.filename = filename
        self.__data_workbook = None
        self.__active_table = None
        self.__location_list = location_list
        self.logger = Logger(logger_name=__name__).get_logger()

    @property
    def get_workbook(self):
        return self.__data_workbook

    @property
    def get_active_table(self):
        return self.__active_table
    
    @property
    def get_location_list(self):
        return self.__location_list

    @staticmethod
    def to_xlsx(workbook, dataframe, data_type: str):
        try:
            if (isinstance(data_type, str)):
                # Creates a new sheet in xlsx file with name = data_type
                if (data_type == 'forecast'):
                    if ('forecast' in workbook.sheetnames):
                        pass
                    else:
                        sheet = workbook.create_sheet('forecast')

                    sheet = workbook['forecast']

                elif(data_type == 'observation'):
                    if ('observation' in workbook.sheetnames):
                        pass
                    else:
                        sheet = workbook.create_sheet('observation')

                    sheet = workbook['observation']
            else:
                raise ValueError(f'Error while creating sheet inside workbook')

            for row in dataframe_to_rows(dataframe, index=False, header=True):
                sheet.append(row)

            return workbook

        except Exception as error:
            raise error
    
    def add_drop_down(self, position:str):
        data_val = DataValidation(type='list', formula1=position)
        self.__active_table.add_data_validation(data_val)
        return data_val

    def set_date_cell(self, date: str, pos_in_table: str):
        check_d_type(pos_in_table,str)
        check_d_type(pos_in_table, str)
        
        date = datetime.datetime.strptime(date, '%Y-%m-%d')
        self.__active_table[pos_in_table] = date                
    
    def set_cell(self, data_list :list, pos_in_table_list: list):
        check_d_type(pos_in_table_list,list)
        check_d_type(data_list, list)

        try:
            for idx, item in enumerate(zip(pos_in_table_list, data_list)):
                pos_in_table = str(item[0])
                data = item[1]            
                self.__active_table[pos_in_table] = data

        except Exception as general_error:
            self.logger.error(f'Error: {general_error}')
        
    def set_active_table(self, table_name: str):
        check_d_type(table_name, str)
        if (table_name in self.__data_workbook.sheetnames):
            pass

        else:
            self.__data_workbook.create_sheet(table_name) 
                   
        self.__active_table = self.__data_workbook[table_name]             

    def load_sheet(self, **kwargs) -> None:
        try:
            self.logger.info(f'Trying to read sheet at: {self.filename}')
            if (check_exists(filename=self.filename) == False):
                raise OSError

            if (kwargs.get('keep_vba') is not None):
                workbook = load_workbook(self.filename, keep_vba=kwargs.pop('keep_vba'), keep_links=True)

            else:
                workbook = load_workbook(self.filename, keep_links=True)

            self.__data_workbook = workbook
            self.logger.info(f'Done.')

        except OSError as os_error:
            self.logger.error(f'OS Error: {os_error}')
            
        except Exception as general_error:
            self.logger.error(f'Error: {general_error}')
            
    def save_sheet(self, workbook, out_file: str):
        try:
            check_d_type(out_file, str)
            path = os.path.dirname(out_file)
            self.logger.info(f'Saving file at: {path}')

            if (check_exists(dir_path=path) == False):
                self.logger.warning(
                    f'Folder specified to save file does not exits, creating it.')
                os.makedirs(path)
                self.logger.warning(f'Done.')

            workbook.save(out_file)

        except PermissionError as permission_error:
            self.logger.error(f'Permission Error: {permission_error}. The file you want to save is open, please close it.')
            
        except OSError as os_error:
            self.logger.error(f'OS Error: {os_error}')
            
        except Exception as general_error:
            self.logger.error(f'Error: {general_error}')
            
