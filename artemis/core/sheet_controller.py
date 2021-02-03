import datetime
import os
from dataclasses import dataclass

import utils
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class Sheet:
    def __init__(self, filename: str) -> None:
        self.filename = filename
        self.data_sheet = None

    def map_sheet(self, min_row, max_row, min_col, max_col):

        for row in self.data_sheet.iter_rows(min_row=min_row,
                                             max_row = max_row, 
                                             min_col = min_col,
                                             max_col = max_col,
                                             values_only=True):
            data = 

      

    @staticmethod
    def to_xlsx(workbook, dataframe, data_type: str):
        try:
            if (isinstance(data_type, str)):
                # Creates a new sheet in xlsx file with name = data_type
                if (data_type == 'forecast'):
                    if ('forecast' in workbook.sheetnames):
                        pass
                    else:
                        sheet = workbook.create_sheet('forecast')

                    sheet = workbook['forecast']

                elif(data_type == 'observation'):
                    if ('observation' in workbook.sheetnames):
                        pass
                    else:
                        sheet = workbook.create_sheet('observation')

                    sheet = workbook['observation']
            else:
                raise ValueError(f'Error while creating sheet inside workbook')

            for row in dataframe_to_rows(dataframe, index=False, header=True):
                sheet.append(row)

            return workbook

        except Exception as error:
            raise error


    def load_sheet(self, **kwargs) -> None:
        try:
            self.logger.info(f'Trying to read sheet at: {self.filename}')
            if (utils.check_exists(filename=self.filename) == False):
                raise OSError

            if (kwargs.get('keep_vba') is not None):
                workbook = load_workbook(self.filename, keep_vba=kwargs.pop(
                    'keep_vba'), data_only=kwargs.pop('data_only'))

            else:
                workbook = load_workbook(
                    self.filename, data_only=kwargs.pop('data_only'))

            self.data_sheet = workbook
            self.logger.info(f'Done.')

        except OSError as os_error:
            self.logger.error(f'OS Error: {os_error}')
            raise os_error

        except Exception as error:
            self.logger.error(f'Error: {error}')
            raise error


    def save_sheet(self, workbook, out_file: str):
        try:
            path = os.path.dirname(out_file)
            self.logger.info(f'Saving file at: {path}')

            if (utils.check_exists(dir_path=path) == False):
                self.logger.warning(
                    f'Folder specified to save file does not exits, creating it.')
                os.makedirs(path)
                self.logger.warning(f'Done.')

            workbook.save(out_file)

        except PermissionError as permission_error:
            self.logger.error(
                f'Permission Error: {permission_error}. The file you want to save is open, please close it.')
            raise permission_error

        except OSError as os_error:
            self.logger.error(f'OS Error: {os_error}')
            raise os_error

        except Exception as error:
            self.logger.error(f'Error: {error}')
            raise error
